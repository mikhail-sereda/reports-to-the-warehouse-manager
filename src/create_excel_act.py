import datetime
import os

from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Side, Border, Font
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet


def open_excel_template_openpyxl(file_path: Path) -> Workbook:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    return workbook


def create_worksheets(open_file: Workbook, worksheets_name: list):
    for name in worksheets_name:
        new_worksheet = open_file.copy_worksheet(open_file.worksheets[0])
        new_worksheet.title = name


def get_invoice_number(text: str) -> str:
    text_list = text.split()
    return text_list[2]


def get_invoice_date(text: str) -> str:
    text_list = text.split()
    date = " ".join(text_list[-3:])
    return date


def pars_invoice(file_path: Path):
    invoice_workbook = open_excel_template_openpyxl(file_path)
    work_worksheet = invoice_workbook.worksheets[0]
    invoice_rows = work_worksheet.values
    data_invoice = {}
    invoice_number = ""

    for row in invoice_rows:
        if any(row):
            for index, cell in enumerate(row):
                if cell and "Требование-накладная" in str(cell):
                    invoice_number = get_invoice_number(cell)
                    date_invoice = get_invoice_date(cell)
                    data_invoice[invoice_number] = {"date_invoice": date_invoice}
                    data_invoice[invoice_number] = {"path": file_path.parent}
                    break
                if cell and "Подразделение" in str(cell):
                    data_invoice[invoice_number]["department"] = row[index + 3]
                    break
                if isinstance(cell, (int, float)):
                    material = row[2]  # столбец B
                    quantity = row[11]  # столбец K
                    unit = row[12]
                    data_invoice[invoice_number].setdefault("materials", list()).append(
                        (material, quantity, unit)
                    )

                    break
    invoice_workbook.close()
    return data_invoice


def create_act(
    invoice_data: dict[dict],
    position1: str,
    name1: str,
    position2: str,
    name2: str,
    reason: str,
    path_template: Path = "static/templates/Шаблон Акт на списание.xlsx",
):
    invoice_workbook = open_excel_template_openpyxl(path_template)
    work_worksheet = invoice_workbook.worksheets[0]

    thins = Side(border_style="thin", color="000000")

    invoice_number = list(invoice_data.keys())[0]
    date_now = datetime.datetime.now()
    date_yesterday = date_now - datetime.timedelta(days=1)

    # заполняем шапку документа
    fill_act_header(
        date_now, invoice_number, name1, name2, position1, position2, work_worksheet
    )

    # заполняем подписи документа и временно копируем
    temp_footer_row = fill_and_get_copy_act_footer(
        name1, name2, position1, position2, work_worksheet
    )

    # удаляем всё что ниже таблицы и выключаем объединение ячеек
    clear_cells_below_row(row_to_clear_below=26, work_worksheet=work_worksheet)

    # заполняем временный список кортежем(строками) со значениями каждой ячейки
    data_for_write: list = prepare_data_for_write(
        date_yesterday, invoice_data, invoice_number, reason
    )
    # добавляем сформированные строки на лист
    for row in data_for_write:
        work_worksheet.append(row)
        added_row_index: int = work_worksheet.max_row  # последняя заполненная строка
        for index, cell in enumerate(work_worksheet[added_row_index]):
            # рисуем границы ячеек (нужно нарисовать до объединения) для объединяемых ячеек границы только для верхней левой
            if index in [0, 1, 4, 5, 7, 8, 10]:
                cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)
                cell.font = Font(name="Times New Roman", size=11)

    # добавляем ранее сохраненный блок подписей
    for row2 in temp_footer_row:
        work_worksheet.append(row2)

    # стилизуем таблицу
    style_and_format_table(thins, work_worksheet)

    # сохраняем сформированный документ в ту же папку, что и исходный документ
    invoice_workbook.save(
        os.path.join(
            invoice_data[invoice_number]["path"],
            f"Акт на списание {invoice_number} {invoice_data[invoice_number]['department']} {date_now.strftime('%d-%m-%Y')}.xlsx",
        )
    )


def fill_act_header(
    date_now: datetime.datetime,
    invoice_number: str,
    name1: str,
    name2: str,
    position1: str,
    position2: str,
    work_worksheet: Worksheet,
):
    cells_data = [
        {"row": 10, "column": 4, "value": invoice_number},
        {"row": 16, "column": 5, "value": position1},
        {"row": 16, "column": 8, "value": name1},
        {"row": 19, "column": 5, "value": position2},
        {"row": 19, "column": 8, "value": name2},
        {
            "row": 22,
            "column": 2,
            "value": f"Составила настоящий акт о том, что произошла порча (потеря) товаров и готовой продукции "
            f"{date_now.strftime('%d.%m.%Y')} подлежащее списанию по решению руководителя.",
        },
    ]
    for data in cells_data:
        work_worksheet.cell(**data)


def fill_and_get_copy_act_footer(
    name1: str, name2: str, position1: str, position2: str, work_worksheet: Worksheet
) -> list[tuple]:
    work_worksheet.cell(row=28, column=5, value=position1)
    work_worksheet.cell(row=28, column=10, value=name1)
    work_worksheet.cell(row=30, column=5, value=position2)
    work_worksheet.cell(row=30, column=10, value=name2)
    return list(work_worksheet.values)[-6:]  # временно копирую нижнюю часть


def clear_cells_below_row(row_to_clear_below: int, work_worksheet: Worksheet):
    for merged_range in list(work_worksheet.merged_cells.ranges):
        # Получаем границы объединенного диапазона
        min_col, min_row, max_col, max_row = merged_range.bounds

        # Если нижняя граница объединения больше или равна указанной строке
        if min_row >= row_to_clear_below:
            work_worksheet.unmerge_cells(str(merged_range))
    # Удаляем все строки ниже указанной (удаляем всегда 26 строку пока максимальная заполненная строка не станет 25)
    while work_worksheet.max_row >= row_to_clear_below:
        work_worksheet.delete_rows(row_to_clear_below)


def prepare_data_for_write(
    date_yesterday: datetime.datetime,
    invoice_data: dict,
    invoice_number: str,
    reason: str,
) -> list[tuple]:
    data_for_write = []
    for index, material in enumerate(
        invoice_data[invoice_number]["materials"], start=1
    ):
        if index <= len(invoice_data[invoice_number]["materials"]):
            data_for_write.append(
                (
                    index,
                    material[0],
                    None,
                    None,
                    material[1],
                    material[2],
                    None,
                    date_yesterday.strftime("%d.%m.%Y"),
                    invoice_data[invoice_number]["department"],
                    None,
                    reason,
                )
            )
    return data_for_write


def style_and_format_table(border_style: Side, work_worksheet: Worksheet):
    for i, cells in enumerate(work_worksheet, start=1):
        cells_value = [cell.value for cell in cells]
        if isinstance(cells_value[0], int):
            for start_col, end_col in (
                (2, 4),
                (6, 7),
                (9, 10),
            ):  # Объединение ячеек в таблице
                work_worksheet.merge_cells(
                    start_row=i, start_column=start_col, end_row=i, end_column=end_col
                )

            cells[1].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cells[8].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            for j in [0, 4, 5, 6, 7, 10]:
                cells[j].alignment = Alignment(horizontal="center", vertical="center")

            work_worksheet.row_dimensions[i].height = 40

        elif "Члены комиссии:" in cells_value:
            work_worksheet.merge_cells(
                start_row=i, start_column=2, end_row=i, end_column=4
            )
            work_worksheet.merge_cells(
                start_row=i, start_column=5, end_row=i, end_column=6
            )
            cells[1].alignment = Alignment(horizontal="right", vertical="center")
            cells[1].font = Font(name="Times New Roman", size=12)
            for j in range(4, 11):
                cells[j].alignment = Alignment(horizontal="center", vertical="center")
                cells[j].font = Font(name="Times New Roman", size=12)

        elif "(должность)" in cells_value:
            work_worksheet.merge_cells(
                start_row=i, start_column=5, end_row=i, end_column=6
            )
            work_worksheet.merge_cells(
                start_row=i + 1, start_column=5, end_row=i + 1, end_column=6
            )

            for j in range(4, 11):
                cells[j].alignment = Alignment(horizontal="center", vertical="center")
                work_worksheet.row_dimensions[i + 1].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                cells[j].font = Font(name="Times New Roman", size=12)
                work_worksheet.row_dimensions[i + 1].font = Font(
                    name="Times New Roman", size=12
                )
            cells[4].border = Border(top=border_style)
            cells[7].border = Border(top=border_style)
            cells[9].border = Border(top=border_style)
    work_worksheet.col_breaks.append(Break(id="11"))
    work_worksheet.page_margins = PageMargins(
        left=0.75, right=0.2, top=0.5, bottom=1, header=0.2, footer=0.2
    )
