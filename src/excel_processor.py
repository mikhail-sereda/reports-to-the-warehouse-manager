from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import xml.etree.ElementTree as ET

# Парсинг из файла
tree = ET.parse(Path("static/data/data.xml"))
root = tree.getroot()

YELLOW_FILL = PatternFill(
    start_color="FFFF00", end_color="FFFF00", fill_type="solid"  # Желтый
)
RED_FILL = PatternFill(
    start_color="FF0000", end_color="FF0000", fill_type="solid"  # Желтый
)


def get_osg(name_product: str):
    for book in root.findall(".//wares"):
        name_pars = book.get("name")
        date = book.get("expirationvalue")
        if (
            name_product.lower() in name_pars.lower()
            or name_pars.lower() in name_product.lower()
        ):
            return date


def iter_excel_openpyxl(file_path: Path, osg: int):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.worksheets[-1]
    rows = sheet.iter_rows()
    report_text = ""

    for row in rows:
        name = str(row[0].value)
        if any(isinstance(cell.value, int) and cell.value != 0 for cell in row):
            expiration_value = get_osg(name)
            if not expiration_value:
                row[0].fill = YELLOW_FILL
                report_text += f"{name.replace(', кг', '')}:{expiration_value} - н.д.\n"
                continue
            for cell in row:
                header_cell = sheet.cell(
                    row=1, column=cell.column
                )  # Абсолютный доступ по координатам
                date_header = header_cell.value
                if cell.value is not None and isinstance(date_header, datetime):
                    # Проверяем, является ли значение числом (int или float)
                    if isinstance(cell.value, (int, float)):
                        days_diff = (datetime.now() - date_header).days

                        if int(expiration_value) - days_diff <= int(
                            int(expiration_value) * ((osg - 20) / 100)
                        ):
                            cell.fill = RED_FILL
                            report_text += f"{name.replace(', кг', '')} - "
                            report_text += f"от {date_header.strftime('%d.%m.%y')} - {cell.value} шт., \n"

                        elif int(expiration_value) - days_diff <= int(
                            int(expiration_value) * (osg / 100)
                        ):
                            cell.fill = YELLOW_FILL
                            report_text += f"{name.replace(', кг', '')}: - "
                            report_text += f"от {date_header.strftime('%d.%m.%y')} - {cell.value} шт., \n"

    output_path = Path(f"{file_path.parent}/Отчет {file_path.stem}.xlsx")
    workbook.save(output_path)
    return output_path, report_text
